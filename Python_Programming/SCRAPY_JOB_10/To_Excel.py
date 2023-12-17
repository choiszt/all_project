import openpyxl
import csv
excel=openpyxl.load_workbook("E:\pycharm project\python程序设计\python大作业\SCRAPY_JOB_10\大作业.xlsx")
path="E:\pycharm project\python程序设计\python大作业\SCRAPY_JOB_10\\"
if("北邮" not in excel.sheetnames):#初始化标签
    excel["Sheet1"].title='北邮'
    excel.create_sheet("西电")
    excel.create_sheet("成电")
    excel.create_sheet("分类")
def set(sheet):
    sheet.cell(1, 1).value = '序号'
    sheet.cell(1, 2).value = '招聘主题'
    sheet.cell(1, 3).value = '发布日期'
    sheet.cell(1, 4).value = '浏览次数'
def classify(name):
    if '大学' in name or '学院' in name or '教' in name or '实验室' in name or '博士' in name or '研究院' in name:
        return '教育企业'
    elif '科技' in name or '研发' in name or '信息' in name or '网络' in name or '计算机' in name or '软件' in name or '数据' in name or '智能' in name:
        return '互联网企业'
    elif '律师' in name or '法' in name or '政' in name:
        return '法律企业'
    elif '通信' in name or '通讯' in name:
        return '通信企业'
    elif '医' in name or '药' in name:
        return '医疗企业'
    elif '芯' in name or '电子' in name or '半导体' in name:
        return '电子相关企业'
    elif '移动' in name or '联通' in name:
        return '电信运营商'
    elif '技术' in name:
        return '互联网企业'
    elif '环' in name:
        return '环保相关企业'
    elif '工业' in name or '航' in name or '工程' in name or '机' in name or '天' in name or '电' in name or '厂' in name or '测' in name or '自动' in name or '控制' in name:
        return '工业、航天企业'
    elif '车' in name or '汽' in name:
        return '汽车企业'
    elif '会计' in name or '工商' in name or '管理' in name:
        return '工商管理企业'
    elif '银行' in name or '证券' in name or '金' in name or '财' in name or '信' in name or '期货' in name or '股' in name or '投资' in name or '经济' in name:
        return '金融企业'
    elif '建' in name:
        return '土木、建设类企业'
    else:
        return '其他企业'
xidianemployer={'教育企业':0,'互联网企业':0,'法律企业':0,'通信企业':0,'医疗企业':0,'电子相关企业':0,'电信运营商':0,'互联网企业':0,'环保相关企业':0,'工业、航天企业':0,'汽车企业':0,'工商管理企业':0,'金融企业':0,'土木、建设类企业':0,"其他企业":0}
chengdianemployer={'教育企业':0,'互联网企业':0,'法律企业':0,'通信企业':0,'医疗企业':0,'电子相关企业':0,'电信运营商':0,'互联网企业':0,'环保相关企业':0,'工业、航天企业':0,'汽车企业':0,'工商管理企业':0,'金融企业':0,'土木、建设类企业':0,"其他企业":0}
buptemployer={'教育企业':0,'互联网企业':0,'法律企业':0,'通信企业':0,'医疗企业':0,'电子相关企业':0,'电信运营商':0,'互联网企业':0,'环保相关企业':0,'工业、航天企业':0,'汽车企业':0,'工商管理企业':0,'金融企业':0,'土木、建设类企业':0,"其他企业":0}
sheet=excel['西电']
set(sheet)
xidianfile=open(path+"XIDIAN_1.csv",'r',encoding='utf-8')
xidianreader=list(csv.reader(xidianfile))
titlelist=[]
xidianlist=[]
xidiandict={}
finaldict={}
for i in range(1,len(xidianreader)):
    sheet.cell(i+1,1).value=i
    sheet.cell(i+1,2).value=xidianreader[i][0]
    if sheet.cell(i+1,2).value not in titlelist:
        titlelist.append(sheet.cell(i+1,2).value)
        xidianlist.append(sheet.cell(i+1,2).value)#去重并存储西电的title
    sheet.cell(i+1,3).value=xidianreader[i][1]
    sheet.cell(i+1,4).value=int(xidianreader[i][2])
    xidiandict[xidianreader[i][0]]=int(xidianreader[i][2])
    finaldict[xidianreader[i][0]]=int(xidianreader[i][2])
processed_xidiandict=sorted(xidiandict.items(),key= lambda x:x[1],reverse=True)
for i in xidianlist:
    xidianemployer[classify(i)]+=1
sheet=excel['成电']
set(sheet)
chengdianfile=open(path+"CHENGDIAN_1.csv",'r',encoding='utf-8')
chengdianreader=list(csv.reader(chengdianfile))
chengdiandict={}
chengdianlist=[]
for i in range(1,len(chengdianreader)):
    sheet.cell(i+1,1).value=i
    sheet.cell(i+1,2).value=chengdianreader[i][0]
    if sheet.cell(i+1,2).value not in titlelist:
        titlelist.append(sheet.cell(i+1,2).value)
        chengdianlist.append(sheet.cell(i+1,2).value)
    sheet.cell(i+1,3).value=chengdianreader[i][1]
    sheet.cell(i+1,4).value=int(chengdianreader[i][2])
    chengdiandict[chengdianreader[i][0]]=int(chengdianreader[i][2])
    finaldict[chengdianreader[i][0]]=int(chengdianreader[i][2])
processed_chengdiandict=sorted(chengdiandict.items(),key=lambda x:x[1],reverse=True)
for i in chengdianlist:
    chengdianemployer[classify(i)]+=1
sheet=excel['北邮']
BUPTfile=open(path+"BEIYOU_1.csv",'r',encoding='utf-8')
BUPTreader=list(csv.reader(BUPTfile))
buptdict={}
buptlist=[]
positiondict={}
for i in range(1,len(BUPTreader)):
    sheet.cell(i+1,1).value=i
    sheet.cell(i+1,2).value=BUPTreader[i][0]
    if sheet.cell(i+1,2).value not in titlelist:
        titlelist.append(sheet.cell(i+1,2).value)
        buptlist.append(sheet.cell(i+1,2).value)
    sheet.cell(i+1,3).value=BUPTreader[i][1]
    sheet.cell(i+1,4).value=int(BUPTreader[i][2])
    sheet.cell(i+1,5).value=int(BUPTreader[i][3])
    positiondict[BUPTreader[i][0]]=int(BUPTreader[i][3])
    buptdict[BUPTreader[i][0]]=int(BUPTreader[i][2])
    finaldict[BUPTreader[i][0]]=int(BUPTreader[i][2])
processed_buptdict=sorted(buptdict.items(),key=lambda x:x[1],reverse=True)
processed_positiondict=sorted(positiondict.items(),key=lambda x:x[1],reverse=True)
for i in buptlist:
    buptemployer[classify(i)]+=1
sheet.cell(1, 1).value = '序号'
sheet.cell(1, 2).value = '招聘主题'
sheet.cell(1, 3).value = '发布日期'
sheet.cell(1, 4).value = '浏览次数'
sheet.cell(1, 5).value = '职位个数'
sheet = excel['分类']
sheet.cell(1, 1).value = '序号'
sheet.cell(1, 2).value = '招聘主题'
sheet.cell(1, 3).value = '雇主类型'
for i in range(len(titlelist)):
    sheet.cell(i+2,1).value=i+1
    sheet.cell(i+2,2).value=titlelist[i]
    sheet.cell(i+2,3).value=classify(titlelist[i])
excel.save("E:\pycharm project\python程序设计\python大作业\SCRAPY_JOB_10\大作业.xlsx")
excel.close()

result=openpyxl.load_workbook("E:\pycharm project\python程序设计\python大作业\SCRAPY_JOB_10\最终结果.xlsx")
if("result" not in result.sheetnames):#初始化标签
    result["Sheet1"].title='result'
wb=result["result"]
wb.cell(1,2).value= '最受北邮学生关注的招聘TOP20'
for i in range(1,21):
    wb.cell(i+1,1).value=i
    wb.cell(i+1,2).value=processed_buptdict[i-1][0]

wb.cell(1, 7).value = '最受西电学生关注的招聘TOP20'
for i in range(1,21):
    wb.cell(i+1,7).value=processed_xidiandict[i-1][0]
wb.cell(1, 12).value = '最受成电学生关注的招聘TOP20'
for i in range(1,21):
    wb.cell(i+1,12).value=processed_chengdiandict[i-1][0]

sorted_chengdian_employer=sorted(chengdianemployer.items(),key=lambda x:x[1],reverse=True)
sorted_xidian_employer=sorted(xidianemployer.items(),key=lambda x:x[1],reverse=True)
sorted_bupt_employer=sorted(buptemployer.items(),key=lambda x:x[1],reverse=True)
wb.cell(23, 2).value = '最受北邮学生关注的雇主类型TOP10'
wb.cell(23, 7).value = '最受西电学生关注的雇主类型TOP10'
wb.cell(23, 12).value = '最受成电学生关注的雇主类型TOP10'
for i in range(1,11):
    wb.cell(23+i,1).value=i
    wb.cell(23+i,2).value=sorted_bupt_employer[i-1][0]
    wb.cell(23+i,7).value=sorted_xidian_employer[i-1][0]
    wb.cell(23+i,12).value=sorted_chengdian_employer[i-1][0]

wb.cell(35, 2).value = '北邮单个公司招聘职位数TOP10'
wb.cell(35, 7).value = '北邮招聘职位总数及对应雇主类型TOP10'
wb.cell(35, 12).value = '职位总数'
for i in range(1, 11):
    wb.cell(35 + i, 1).value = i
    wb.cell(35 + i, 2).value = processed_positiondict[i-1][0]
    wb.cell(35 + i, 7).value =classify(processed_positiondict[i-1][0])
    wb.cell(35 + i, 12).value =processed_positiondict[i-1][1]
processed_finaldict=sorted(finaldict.items(),key=lambda x:x[1],reverse=True)
wb.cell(47, 2).value = '最关注ICT行业的招聘主题TOP10'
wb.cell(47, 7).value = '浏览次数'
for i in range(1, 11):
    wb.cell(47 + i, 1).value = i
    wb.cell(47 + i, 2).value = processed_finaldict[i-1][0]
    wb.cell(47 + i, 7).value = processed_finaldict[i-1][1]
result.save("E:\pycharm project\python程序设计\python大作业\SCRAPY_JOB_10\最终结果.xlsx")